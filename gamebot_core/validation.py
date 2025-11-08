"""Lightweight bronze layer data quality checks."""

from __future__ import annotations

import json
from collections import defaultdict
import logging
import re
import sys
from dataclasses import dataclass
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import numpy as np
import pandas as pd

try:  # Optional dependency for metadata lookups
    import requests  # type: ignore
except Exception:  # pragma: no cover - requests may be unavailable in some envs
    requests = None

base_dir = Path(__file__).resolve().parent.parent
if str(base_dir) not in sys.path:
    sys.path.append(str(base_dir))

import params  # noqa: E402
from .log_utils import get_run_log_dir  # noqa: E402

logger = logging.getLogger(__name__)

SECTION_DESCRIPTIONS: Dict[str, str] = {
    "Rule Checks": "Configured row-level validation rules and their status.",
    "Unique Constraint": "Columns expected to remain unique; duplicates are removed or flagged before load.",
    "Foreign Key Checks": "Ensures values exist in the referenced table; failures show sample offending rows.",
    "Null Value Summary": "Columns containing nulls (non-zero counts only).",
    "Remediation Events": "Automatic data adjustments performed during preprocessing (dedupes, fixes, etc.).",
    "Column Types": "Comparison between pandas data types and database column types.",
    "Notes": "Additional context when no issues were detected.",
    "Upstream Dataset Coverage": "Comparison of survivoR upstream datasets with the datasets ingested during this run.",
    "Unexpected Columns": "Columns discovered in the upstream data that are not present in the warehouse schema.",
    "Missing Columns": "Columns expected by the warehouse schema that were absent in the upstream data.",
}

CURRENT_VALIDATION_SUBDIR: Optional[Path] = None
CURRENT_RUN_ID: Optional[str] = None
CURRENT_RUN_LABEL: Optional[str] = None
GLOBAL_VERSION_SEASONS: Set[str] = set()
GLOBAL_CONFIGURED_DATASETS: Set[str] = set()
GLOBAL_LOADED_DATASETS: Set[str] = set()
GLOBAL_DATASET_METADATA: Dict[str, Dict[str, Any]] = {}
_UPSTREAM_DATASETS_CACHE: Optional[Set[str]] = None
_UPSTREAM_DATASETS_ERROR: Optional[str] = None

ISSUE_LABELS: Dict[str, str] = {
    "rows_dropped_multi_castaway": "Removed rows with multiple castaway ids",
    "multi_target_advantage_split": "Split advantage rows with multiple targets",
    "multi_holder_advantage_split": "Split advantage rows with multiple holders",
    "deduplicated_rows": "Removed duplicate records",
    "challenge_description_stub_creation": "Created challenge description stub rows",
    "invalid_advantage_targets": "Dropped invalid advantage targets",
    "value_coercion": "Values coerced to match schema types",
    "challenge_id_known_fix": "Corrected challenge ids using known fixups",
    "challenge_id_remediation": "Corrected challenge ids using stage-of-game",
    "rows_dropped_missing_castaway_id": "Removed rows missing castaway id",
    "null_castaway_ids": "Rows retain null castaway id",
}


def _short_run_label(run_identifier: str) -> str:
    cleaned = "".join(ch for ch in run_identifier if ch.isalnum())
    if len(cleaned) >= 12:
        return f"{cleaned[:8]}-{cleaned[-4:]}".upper()
    return cleaned.upper()


def _find_existing_run_dir(run_identifier: str) -> Optional[Path]:
    if not VALIDATION_DIR.exists():
        return None
    for child in VALIDATION_DIR.iterdir():
        if not child.is_dir():
            continue
        marker_file = child / ".run_id"
        if marker_file.exists() and marker_file.read_text().strip() == run_identifier:
            return child
    return None


def _next_run_sequence() -> int:
    max_seq = 0
    if VALIDATION_DIR.exists():
        for child in VALIDATION_DIR.iterdir():
            if not child.is_dir():
                continue
            match = RUN_FOLDER_PATTERN.match(child.name)
            if not match:
                continue
            try:
                value = int(match.group(1))
            except ValueError:
                continue
            max_seq = max(max_seq, value)
    return max_seq + 1


def set_validation_run(run_identifier: Optional[str]) -> None:
    """Configure the validation output directory for the active run."""
    global CURRENT_VALIDATION_SUBDIR, CURRENT_RUN_ID, CURRENT_RUN_LABEL
    CURRENT_RUN_ID = run_identifier
    if not run_identifier:
        CURRENT_VALIDATION_SUBDIR = None
        CURRENT_RUN_LABEL = None
        return

    existing_dir = _find_existing_run_dir(run_identifier)
    label = _short_run_label(run_identifier)

    if existing_dir is not None:
        subdir = existing_dir
    else:
        sequence = _next_run_sequence()
        folder_base = f"Run {sequence:04d} - {label} Validation Files"
        subdir = VALIDATION_DIR / folder_base
        counter = 2
        while subdir.exists():
            folder_base = f"Run {sequence:04d} - {label} Validation Files ({counter})"
            subdir = VALIDATION_DIR / folder_base
            counter += 1

    try:
        subdir.mkdir(parents=True, exist_ok=True)
        (subdir / ".run_id").write_text(run_identifier)
        CURRENT_VALIDATION_SUBDIR = subdir
    except PermissionError as e:
        # If we can't create the directory due to permissions, fall back to temp directory
        import tempfile
        from pathlib import Path

        temp_dir = Path(tempfile.gettempdir()) / "gamebot_validation" / folder_base
        try:
            temp_dir.mkdir(parents=True, exist_ok=True)
            (temp_dir / ".run_id").write_text(run_identifier)
            CURRENT_VALIDATION_SUBDIR = temp_dir
            logger = logging.getLogger(__name__)
            logger.warning(
                f"Could not create validation directory {subdir}: {e}. Using temp directory: {temp_dir}"
            )
        except PermissionError:
            # If even temp directory fails due to permissions, disable validation reports
            CURRENT_VALIDATION_SUBDIR = None
            logger = logging.getLogger(__name__)
            logger.warning(
                "Could not create validation directories due to permissions. Validation reports disabled."
            )
    CURRENT_RUN_LABEL = label


def clear_validation_run() -> None:
    """Reset the validation output directory context."""
    global CURRENT_VALIDATION_SUBDIR, CURRENT_RUN_ID, CURRENT_RUN_LABEL
    CURRENT_VALIDATION_SUBDIR = None
    CURRENT_RUN_ID = None
    CURRENT_RUN_LABEL = None


RUN_LOG_DIR = get_run_log_dir()
VALIDATION_DIR = RUN_LOG_DIR / "validation"
CHECK_PATTERN = re.compile(
    r"^(?P<metric>missing_count|duplicate_count)\((?P<column>[^)]+)\)\s*=\s*(?P<expected>-?\d+)$"
)
RUN_FOLDER_PATTERN = re.compile(r"^Run\s+(\d+)\b")


@dataclass(frozen=True)
class ForeignKeyRule:
    target_columns: Tuple[str, ...]
    reference_dataset: str
    reference_columns: Tuple[str, ...]
    allow_null: bool = True


# Cache minimal reference frames so inter-dataset validations can reuse them.
REFERENCE_CACHE: Dict[str, pd.DataFrame] = {}
VALIDATION_SUMMARIES: Dict[str, Dict[str, Any]] = {}
DATA_ISSUES: List[Dict[str, Any]] = []


def register_data_issue(dataset: str, issue_type: str, details: Dict[str, Any]) -> None:
    """Record a remediation or anomaly detected during preprocessing."""
    DATA_ISSUES.append(
        {
            "dataset": dataset,
            "issue_type": issue_type,
            "details": details,
            "timestamp": datetime.utcnow().isoformat(),
        }
    )


def _write_result(dataset_name: str, result: Dict) -> Path:
    """Persist the validation results for observability."""
    output_dir = CURRENT_VALIDATION_SUBDIR or VALIDATION_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"validation_{dataset_name}_{timestamp}.json"
    output_path.write_text(json.dumps(result, indent=2, default=str))
    return output_path


def _evaluate_check(df: pd.DataFrame, check: str) -> Tuple[bool, Dict]:
    """Evaluate a single validation rule against the dataframe."""
    match = CHECK_PATTERN.match(check.strip())
    if not match:
        raise ValueError(f"Unsupported validation rule syntax: '{check}'")

    metric = match.group("metric")
    column = match.group("column")
    expected = int(match.group("expected"))

    if column not in df.columns:
        return False, {
            "check": check,
            "status": "failed",
            "reason": "column_missing",
            "detail": f"Column '{column}' not present in dataframe.",
        }

    if metric == "missing_count":
        observed = int(df[column].isna().sum())
    elif metric == "duplicate_count":
        observed = int(df[column].duplicated().sum())
    else:
        raise ValueError(f"Unsupported validation metric: '{metric}'")

    passed = observed == expected
    result = {
        "check": check,
        "status": "passed" if passed else "failed",
        "observed": observed,
        "expected": expected,
    }
    return passed, result


def _null_count_summary(df: pd.DataFrame) -> Dict[str, int]:
    """Return a mapping of columns to null counts (excluding zeroes)."""
    return {
        column: int(count)
        for column, count in df.isna().sum().items()
        if int(count) > 0
    }


def _run_dataframe_checks(
    dataset_name: str, df: pd.DataFrame, checks: Iterable[str]
) -> Dict[str, Any]:
    """Execute validation rules and return a summary payload."""
    check_results: List[Dict[str, Any]] = []

    for check in checks:
        passed, result = _evaluate_check(df, check)
        check_results.append(result)

    null_summary = _null_count_summary(df)

    if null_summary:
        logger.info(
            "Null counts for %s: %s",
            dataset_name,
            {k: v for k, v in null_summary.items() if v > 0},
        )

    summary: Dict[str, Any] = {
        "dataset": dataset_name,
        "total_checks": len(check_results),
        "failed_checks": sum(1 for r in check_results if r["status"] == "failed"),
        "checks": check_results,
        "missing_values": null_summary,
        "timestamp": datetime.utcnow().isoformat(),
    }

    summary["row_count"] = int(len(df))
    summary["pandas_dtypes"] = {
        column: str(dtype) for column, dtype in df.dtypes.items()
    }

    return summary


def _base_checks(key_column: str) -> List[str]:
    return [
        f"missing_count({key_column}) = 0",
        f"duplicate_count({key_column}) = 0",
    ]


DATASET_CHECKS: Dict[str, List[str]] = {
    "castaway_details": _base_checks("castaway_id"),
    "season_summary": _base_checks("version_season"),
    "episodes": [
        "missing_count(version_season) = 0",
        "missing_count(episode) = 0",
    ],
    "advantage_details": [
        "missing_count(advantage_id) = 0",
        "missing_count(version_season) = 0",
    ],
    "challenge_description": [
        "missing_count(challenge_id) = 0",
        "missing_count(version_season) = 0",
    ],
    "challenge_summary": [
        "missing_count(version_season) = 0",
        "missing_count(challenge_id) = 0",
    ],
    "auction_details": [
        "missing_count(version_season) = 0",
        "missing_count(auction_num) = 0",
        "missing_count(item) = 0",
    ],
    "advantage_movement": [
        "missing_count(version_season) = 0",
        "missing_count(advantage_id) = 0",
        "missing_count(sequence_id) = 0",
    ],
    "journeys": [
        "missing_count(version_season) = 0",
        "missing_count(sog_id) = 0",
    ],
    "tribe_mapping": [
        "missing_count(version_season) = 0",
        "missing_count(tribe) = 0",
    ],
    "vote_history": [
        "missing_count(version_season) = 0",
    ],
}


def validate_bronze_dataset(
    dataset_name: str,
    df: pd.DataFrame,
    db_schema: Optional[Dict[str, str]] = None,
) -> None:
    """Run dataframe-based validations for a given bronze dataset."""
    checks = DATASET_CHECKS.get(dataset_name)
    df_copy = df.copy()

    if checks:
        summary = _run_dataframe_checks(dataset_name, df_copy, checks)
    else:
        null_summary = _null_count_summary(df_copy)
        if null_summary:
            logger.info(
                "Null counts for %s: %s",
                dataset_name,
                {k: v for k, v in null_summary.items() if v > 0},
            )
        summary = {
            "dataset": dataset_name,
            "total_checks": 0,
            "failed_checks": 0,
            "checks": [],
            "missing_values": null_summary,
            "timestamp": datetime.utcnow().isoformat(),
            "notes": [
                "No dataset-specific row checks defined; uniqueness and foreign-key checks still executed."
            ],
        }

    summary.setdefault("row_count", int(len(df_copy)))
    summary.setdefault(
        "pandas_dtypes",
        {column: str(dtype) for column, dtype in df_copy.dtypes.items()},
    )
    summary.setdefault("notes", [])
    if db_schema:
        summary["db_column_types"] = {
            column: db_schema.get(column) for column in db_schema
        }

    global GLOBAL_VERSION_SEASONS
    if "version_season" in df_copy.columns:
        version_values = (
            df_copy["version_season"]
            .dropna()
            .astype(str)
            .str.strip()
            .replace({"": pd.NA})
            .dropna()
        )
        present = sorted(set(version_values.tolist()))
        coverage: Dict[str, Any] = {
            "present": present,
            "present_count": len(present),
        }
        if dataset_name == "season_summary":
            GLOBAL_VERSION_SEASONS = set(present)
            coverage["baseline"] = "season_summary"
        elif GLOBAL_VERSION_SEASONS:
            coverage["baseline"] = "season_summary"
        if GLOBAL_VERSION_SEASONS:
            missing = sorted(GLOBAL_VERSION_SEASONS - set(present))
            unexpected = sorted(set(present) - GLOBAL_VERSION_SEASONS)
            if missing:
                coverage["missing"] = missing
                coverage["missing_count"] = len(missing)
            if unexpected:
                coverage["unexpected"] = unexpected
                coverage["unexpected_count"] = len(unexpected)
        summary["version_season_coverage"] = coverage

    status = "passed" if summary["failed_checks"] == 0 else "failed"

    unique_columns = _unique_constraint_for_dataset(dataset_name)
    unique_result: Optional[Dict[str, Any]] = None
    nullable_columns = set()
    if unique_columns:
        # Uniqueness is derived from Database/table_config.json, so we surface the
        # configured constraint even when there are no explicit checks.
        unique_result = _check_unique_constraint(dataset_name, df_copy, unique_columns)
        nullable_columns = _nullable_unique_columns(dataset_name)
        unique_result["nullable_columns"] = list(nullable_columns)
        key_null_counts = {
            col: summary["missing_values"].get(col, 0)
            for col in unique_columns
            if summary["missing_values"].get(col, 0)
        }
        if key_null_counts:
            unique_result["key_null_counts"] = key_null_counts
            summary.setdefault("notes", []).append(
                "Unique key columns containing nulls (see Unique Null Samples section): "
                + ", ".join(f"{col}={count}" for col, count in key_null_counts.items())
            )
        summary["unique_constraint"] = unique_result
        if unique_result.get("status") == "failed":
            status = "failed"
    else:
        summary["unique_constraint"] = None

    _register_reference_snapshot(dataset_name, df_copy)
    fk_results = _run_foreign_key_checks(dataset_name, df_copy)
    summary["foreign_keys"] = fk_results
    for entry in fk_results:
        null_count = entry.get("null_count")
        if null_count:
            cols = entry.get("columns")
            if isinstance(cols, (list, tuple)):
                cols_display = ", ".join(map(str, cols))
            else:
                cols_display = str(cols)
            summary.setdefault("notes", []).append(
                f"Foreign key columns {cols_display} contain {null_count} null rows (see Foreign Key Null Samples section)."
            )
    if any(result.get("status") == "failed" for result in fk_results):
        status = "failed"

    # Fold in any remediation messages emitted by preprocessing (e.g., challenge ID corrections).
    _collect_dataset_issues(dataset_name, summary)

    summary["status"] = status
    result_path = _write_result(dataset_name, summary)
    summary["result_path"] = str(result_path)

    VALIDATION_SUMMARIES[dataset_name] = summary

    if status != "passed":
        logger.error("Validation failed for %s. See %s", dataset_name, result_path)
        raise ValueError(
            f"Validation failed for dataset '{dataset_name}' (details in {result_path})"
        )

    logger.info(
        "Validation succeeded for %s (results written to %s)", dataset_name, result_path
    )


def append_dataset_issues(dataset_name: str) -> None:
    summary = VALIDATION_SUMMARIES.get(dataset_name)
    if not summary:
        return
    _collect_dataset_issues(dataset_name, summary)


def _register_reference_snapshot(dataset_name: str, df: pd.DataFrame) -> None:
    """Persist minimal slices of datasets used as foreign-key references."""
    reference_columns_map: Dict[str, Tuple[str, ...]] = {
        "castaway_details": ("castaway_id",),
        "challenge_description": ("version_season", "challenge_id"),
        "challenge_results": ("version_season", "sog_id", "challenge_id"),
        "season_summary": ("version_season",),
    }
    columns = reference_columns_map.get(dataset_name)
    if not columns:
        return
    subset = df[list(columns)].dropna().drop_duplicates()
    REFERENCE_CACHE[dataset_name] = subset


def _table_config_entry(dataset_name: str) -> Optional[Dict[str, Any]]:
    table_name: Optional[str] = None
    for entry in params.dataset_order:
        if entry.get("dataset") == dataset_name:
            table_name = entry.get("table_name")
            break

    if not table_name:
        return None

    for value in params.table_config.values():
        if isinstance(value, dict) and value.get("table_name") == table_name:
            return value
    return None


def _unique_constraint_for_dataset(dataset_name: str) -> Optional[Tuple[str, ...]]:
    entry = _table_config_entry(dataset_name)
    if not entry:
        return None

    columns = entry.get("unique_constraint_columns", [])
    if columns:
        return tuple(columns)
    return None


def _nullable_unique_columns(dataset_name: str) -> Set[str]:
    entry = _table_config_entry(dataset_name)
    if not entry:
        return set()
    nullable = entry.get("unique_constraint_nullable_columns", [])
    return {str(column) for column in nullable}


def _consume_dataset_issues(dataset_name: str) -> List[Dict[str, Any]]:
    matched = [issue for issue in DATA_ISSUES if issue["dataset"] == dataset_name]
    if matched:
        DATA_ISSUES[:] = [issue for issue in DATA_ISSUES if issue not in matched]
    return matched


def _collect_dataset_issues(dataset_name: str, summary: Dict[str, Any]) -> None:
    existing_issues = summary.get("issues") or []
    issues = _consume_dataset_issues(dataset_name)
    if not issues:
        summary["issues"] = existing_issues
        return

    filtered = [
        issue for issue in issues if issue.get("issue_type") != "value_coercion"
    ]
    summary["issues"] = existing_issues + filtered


def register_configured_dataset(dataset_name: str) -> None:
    GLOBAL_CONFIGURED_DATASETS.add(str(dataset_name))


def record_dataset_metadata(
    dataset_name: str,
    table_name: str,
    observed_columns: Iterable[str],
    db_columns: Iterable[str],
    *,
    auto_columns: Iterable[str] = (),
) -> None:
    metadata = GLOBAL_DATASET_METADATA.setdefault(
        dataset_name,
        {
            "table_name": table_name,
            "observed": set(),
            "db_columns": set(db_columns),
        },
    )
    metadata["table_name"] = table_name
    observed_set = metadata.setdefault("observed", set())
    observed_set.update(str(column) for column in observed_columns)
    metadata["db_columns"] = set(str(column) for column in db_columns)
    metadata["auto_columns"] = set(str(column) for column in auto_columns)
    metadata["loaded"] = True
    GLOBAL_LOADED_DATASETS.add(str(dataset_name))


def _github_api_headers() -> Dict[str, str]:
    token = os.getenv("GITHUB_TOKEN")
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def _fetch_upstream_dataset_names() -> Optional[Set[str]]:
    global _UPSTREAM_DATASETS_CACHE, _UPSTREAM_DATASETS_ERROR
    if _UPSTREAM_DATASETS_CACHE is not None:
        return _UPSTREAM_DATASETS_CACHE
    if _UPSTREAM_DATASETS_ERROR is not None:
        return None
    if requests is None:
        _UPSTREAM_DATASETS_ERROR = "requests library not available"
        return None
    url = "https://api.github.com/repos/doehm/survivoR/contents/dev/json"
    try:
        response = requests.get(url, headers=_github_api_headers(), timeout=30)
        response.raise_for_status()
        payload = response.json()
        datasets = {
            str(item.get("name", "")).rsplit(".json", 1)[0]
            for item in payload
            if item.get("name", "").endswith(".json")
        }
        _UPSTREAM_DATASETS_CACHE = {name for name in datasets if name}
        return _UPSTREAM_DATASETS_CACHE
    except Exception as exc:  # pragma: no cover - network variability
        _UPSTREAM_DATASETS_ERROR = str(exc)
        return None


def _check_unique_constraint(
    dataset_name: str, df: pd.DataFrame, columns: Tuple[str, ...]
) -> Dict[str, Any]:
    """Ensure the configured unique constraint columns remain unique."""
    result: Dict[str, Any] = {
        "columns": list(columns),
        "status": "passed",
    }

    if not columns:
        return result

    nullable_columns = _nullable_unique_columns(dataset_name)
    null_samples: Dict[str, List[Dict[str, Any]]] = {}
    non_nullable_columns = [col for col in columns if col not in nullable_columns]

    for column in columns:
        column_mask = df[column].isna()
        if column_mask.any():
            null_samples[column] = (
                df.loc[column_mask, list(columns)].head(10).to_dict("records")
            )

    if null_samples:
        result["null_samples"] = null_samples

    if non_nullable_columns:
        missing_mask = df[non_nullable_columns].isna().any(axis=1)
        if missing_mask.any():
            sample = df.loc[missing_mask, list(columns)].head(10).to_dict("records")
            logger.error(
                "Unique constraint columns %s for %s contain nulls (sample: %s)",
                columns,
                dataset_name,
                sample,
            )
            result["status"] = "failed"
            result["missing_sample"] = sample

    df_for_check = df
    if nullable_columns:
        df_for_check = df_for_check[
            df_for_check[list(nullable_columns)].notna().all(axis=1)
        ]

    if not df_for_check.empty:
        duplicate_mask = df_for_check.duplicated(subset=list(columns), keep=False)
        if duplicate_mask.any():
            sample = (
                df_for_check.loc[duplicate_mask, list(columns)]
                .head(10)
                .to_dict("records")
            )
            logger.error(
                "Duplicate rows detected for %s on columns %s (sample: %s)",
                dataset_name,
                columns,
                sample,
            )
            result["status"] = "failed"
            result["duplicate_sample"] = sample

    if result["status"] == "failed":
        raise ValueError(
            f"Dataset '{dataset_name}' violates uniqueness on columns {columns}"
        )

    return result


FOREIGN_KEY_RULES: Dict[str, List[ForeignKeyRule]] = {
    "vote_history": [
        ForeignKeyRule(
            target_columns=("castaway_id",),
            reference_dataset="castaway_details",
            reference_columns=("castaway_id",),
            allow_null=True,
        ),
        ForeignKeyRule(
            target_columns=("vote_id",),
            reference_dataset="castaway_details",
            reference_columns=("castaway_id",),
            allow_null=True,
        ),
        ForeignKeyRule(
            target_columns=("voted_out_id",),
            reference_dataset="castaway_details",
            reference_columns=("castaway_id",),
            allow_null=True,
        ),
        ForeignKeyRule(
            target_columns=("version_season", "challenge_id"),
            reference_dataset="challenge_description",
            reference_columns=("version_season", "challenge_id"),
            allow_null=True,
        ),
    ],
    "journeys": [
        ForeignKeyRule(
            target_columns=("castaway_id",),
            reference_dataset="castaway_details",
            reference_columns=("castaway_id",),
            allow_null=False,
        ),
        ForeignKeyRule(
            target_columns=("version_season",),
            reference_dataset="season_summary",
            reference_columns=("version_season",),
            allow_null=False,
        ),
    ],
    "castaway_scores": [
        ForeignKeyRule(
            target_columns=("castaway_id",),
            reference_dataset="castaway_details",
            reference_columns=("castaway_id",),
            allow_null=False,
        ),
        ForeignKeyRule(
            target_columns=("version_season",),
            reference_dataset="season_summary",
            reference_columns=("version_season",),
            allow_null=False,
        ),
    ],
}

FK_CONTEXT_COLUMNS: Dict[str, List[str]] = {
    "vote_history": [
        "episode",
        "sog_id",
        "castaway_id",
        "vote_id",
        "voted_out_id",
        "tribe_status",
        "immunity",
        "vote_event",
    ],
    "journeys": [
        "episode",
        "sog_id",
        "castaway_id",
        "lost_vote",
        "reward_details",
    ],
}


def _run_foreign_key_checks(
    dataset_name: str, df: pd.DataFrame
) -> List[Dict[str, Any]]:
    """Verify configured foreign key relationships using cached reference data."""
    results: List[Dict[str, Any]] = []
    rules = FOREIGN_KEY_RULES.get(dataset_name)
    if not rules:
        return results

    for rule in rules:
        result: Dict[str, Any] = {
            "columns": list(rule.target_columns),
            "referenced_table": rule.reference_dataset,
            "reference_columns": list(rule.reference_columns),
            "allow_null": rule.allow_null,
            "status": "skipped",
        }

        reference_df = REFERENCE_CACHE.get(rule.reference_dataset)
        if reference_df is None or reference_df.empty:
            logger.debug(
                "Skipping FK check for %s → %s (reference data unavailable).",
                dataset_name,
                rule.reference_dataset,
            )
            result["reason"] = "reference_missing"
            results.append(result)
            continue

        target_cols = list(rule.target_columns)
        missing_cols = [col for col in target_cols if col not in df.columns]
        if missing_cols:
            logger.debug(
                "Skipping FK check for %s (columns %s missing).",
                dataset_name,
                missing_cols,
            )
            result["reason"] = "target_missing"
            results.append(result)
            continue

        context_cols = [
            col for col in FK_CONTEXT_COLUMNS.get(dataset_name, []) if col in df.columns
        ]

        subset = df[target_cols].copy()
        if not rule.allow_null:
            subset = subset.dropna()
        else:
            null_mask = df[target_cols].isna().any(axis=1)
            if null_mask.any():
                result["null_count"] = int(null_mask.sum())
                result["null_samples"] = (
                    df.loc[null_mask, list(dict.fromkeys(target_cols + context_cols))]
                    .head(10)
                    .to_dict("records")
                )
            subset = subset.dropna()

        if subset.empty:
            result["status"] = "passed"
            result["note"] = "No rows to validate"
            results.append(result)
            continue

        ref_columns = list(rule.reference_columns)
        reference_keys = {
            tuple(_safe_scalar(value) for value in row)
            for row in reference_df[ref_columns].dropna().to_records(index=False)
        }

        unmatched: List[Dict[str, Any]] = []
        seen_unmatched: Set[Tuple[Any, ...]] = set()
        combined_cols = list(dict.fromkeys(target_cols + context_cols))
        working_df = df.loc[subset.index, combined_cols] if combined_cols else None

        for idx, values in zip(subset.index, subset.to_records(index=False)):
            tupled = tuple(_safe_scalar(v) for v in values)
            if any(v is None for v in tupled):
                continue
            if tupled not in reference_keys:
                if tupled in seen_unmatched:
                    continue
                seen_unmatched.add(tupled)
                if working_df is not None:
                    unmatched.append(working_df.loc[idx].to_dict())
                else:
                    unmatched.append(dict(zip(target_cols, tupled)))

        if unmatched:
            logger.warning(
                "Foreign key validation failed for %s columns %s -> %s.%s. Sample invalid keys: %s",
                dataset_name,
                target_cols,
                rule.reference_dataset,
                ref_columns,
                unmatched,
            )
            result["status"] = "failed"
            result["sample_records"] = unmatched[:10]
        else:
            result["status"] = "passed"

        results.append(result)

    return results


def _safe_scalar(value: Optional[object]) -> Optional[object]:
    """Normalize numpy scalar types into native Python for set comparisons."""
    scalar = value
    if isinstance(value, (np.generic,)):
        scalar = value.item()
    if scalar is None:
        return None
    try:
        if pd.isna(scalar):
            return None
    except TypeError:
        pass
    return scalar


def _values_equal(left: Any, right: Any) -> bool:
    if left is right:
        return True
    if isinstance(left, (list, tuple)) and isinstance(right, (list, tuple)):
        if len(left) != len(right):
            return False
        return all(_values_equal(lhs, rhs) for lhs, rhs in zip(left, right))
    try:
        left_null = pd.isna(left)
    except TypeError:
        left_null = False
    try:
        right_null = pd.isna(right)
    except TypeError:
        right_null = False
    if bool(left_null) and bool(right_null):
        return True
    return left == right


def finalise_validation_reports(run_identifier: Optional[str] = None) -> Optional[Path]:
    """Build an Excel workbook summarising validation outcomes for all datasets."""
    if not VALIDATION_SUMMARIES:
        logger.info("No validation summaries captured; skipping Excel export.")
        return None

    label = (
        CURRENT_RUN_LABEL
        or (_short_run_label(run_identifier) if run_identifier else None)
        or "RUN"
    )
    safe_label = re.sub(r"[^A-Za-z0-9_-]", "_", label)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"data_quality_{safe_label}_{timestamp}.xlsx"

    target_dir = CURRENT_VALIDATION_SUBDIR or VALIDATION_DIR
    target_dir.mkdir(parents=True, exist_ok=True)

    workbook_path = target_dir / filename

    try:
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:  # type: ignore[arg-type]
            for dataset, summary in sorted(VALIDATION_SUMMARIES.items()):
                _write_dataset_sheet(writer, dataset, summary)
            _write_metadata_summary_sheet(writer)
        logger.info("Data quality workbook written to %s", workbook_path)
        return workbook_path
    except (ModuleNotFoundError, ValueError) as exc:
        logger.warning(
            "Unable to generate Excel data quality report (%s) using openpyxl engine.",
            exc,
        )
        return None
    finally:
        GLOBAL_VERSION_SEASONS.clear()
        GLOBAL_CONFIGURED_DATASETS.clear()
        GLOBAL_LOADED_DATASETS.clear()
        GLOBAL_DATASET_METADATA.clear()
        global _UPSTREAM_DATASETS_CACHE, _UPSTREAM_DATASETS_ERROR
        _UPSTREAM_DATASETS_CACHE = None
        _UPSTREAM_DATASETS_ERROR = None
        VALIDATION_SUMMARIES.clear()
        DATA_ISSUES.clear()
        if run_identifier:
            clear_validation_run()


def _summarize_issue_details(issue_type: str, details: Dict[str, Any]) -> str:
    parts: List[str] = []
    if "rows_removed" in details:
        before = details.get("before")
        after = details.get("after")
        if before is not None and after is not None:
            parts.append(
                f"Removed {details['rows_removed']} rows (before {before} → after {after})"
            )
        else:
            parts.append(f"Removed {details['rows_removed']} rows")
    if "rows_added" in details:
        parts.append(f"Added {details['rows_added']} rows")
    if issue_type == "challenge_description_stub_creation":
        target_table = details.get("target_table")
        stub_columns = details.get("stub_columns")
        if target_table:
            parts.append(f"Target table: {target_table}")
        if stub_columns:
            parts.append("Stub columns: " + ", ".join(map(str, stub_columns)))
    if "rows_split" in details:
        parts.append(f"Split {details['rows_split']} rows")
    if "rows_affected" in details:
        parts.append(f"Affected {details['rows_affected']} rows")
    if "rows_corrected" in details:
        parts.append(f"Corrected {details['rows_corrected']} rows")
    if issue_type == "castaway_id_backfilled" and details.get("rows_updated"):
        parts.append(f"Backfilled {details['rows_updated']} castaway_id values")
    if issue_type == "castaway_id_fuzzy_backfill":
        src = details.get("source_name")
        matched = details.get("matched_name")
        cid = details.get("castaway_id")
        if src and matched:
            message = f"Fuzzy matched '{src}' → '{matched}'"
            if cid:
                message += f" (castaway_id={cid})"
            parts.append(message)
    if "original_rows" in details:
        parts.append(f"Original rows: {len(details['original_rows'])}")
    if "result_rows" in details:
        parts.append(f"Result rows: {len(details['result_rows'])}")
    if "removed_rows" in details:
        parts.append(f"Removed rows: {len(details['removed_rows'])}")
    if "added_rows" in details:
        parts.append(f"Added rows: {len(details['added_rows'])}")
    if "subset_columns" in details:
        subset = details["subset_columns"]
        parts.append(
            "Unique columns: "
            + (", ".join(subset) if isinstance(subset, list) else str(subset))
        )
    if "distinct_targets" in details:
        parts.append(f"Dropped targets: {', '.join(details['distinct_targets'])}")
    if "sample_vote_events" in details:
        parts.append(f"Sample vote events: {', '.join(details['sample_vote_events'])}")
    if "column" in details and issue_type == "value_coercion":
        parts.append(f"Column coerced: {details['column']}")
    if not parts:
        parts.append(json.dumps(details, default=str, sort_keys=True))
    return "; ".join(parts)


def _extract_detail_records(
    details: Dict[str, Any], remediation_id: Optional[str] = None
) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []

    def _append_rows(key: str, state: str) -> None:
        value = details.get(key)
        if isinstance(value, list):
            for entry in value:
                if isinstance(entry, dict):
                    record = entry.copy()
                    record["record_state"] = state
                    if remediation_id:
                        record["remediation_id"] = remediation_id
                    records.append(record)

    _append_rows("original_rows", "original")
    _append_rows("result_rows", "result")
    _append_rows("removed_rows", "removed")
    _append_rows("added_rows", "added")
    _append_rows("reference_rows", "reference")

    return records


def _expand_remediation_issues(
    issues: List[Dict[str, Any]],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Split remediation issues into summary and detailed record tables."""
    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []

    for index, issue in enumerate(issues, start=1):
        dataset = issue.get("dataset")
        issue_type = issue.get("issue_type", "")
        label = ISSUE_LABELS.get(issue_type, issue_type.replace("_", " ").capitalize())
        timestamp = issue.get("timestamp")
        details = issue.get("details")
        remediation_id = f"{dataset or 'dataset'}:{issue_type or 'issue'}:{index:03d}"
        if isinstance(details, dict):
            summary_rows.append(
                {
                    "dataset": dataset,
                    "issue": label,
                    "timestamp": timestamp,
                    "impact": _summarize_issue_details(issue_type, details),
                    "remediation_id": remediation_id,
                }
            )
            issue_changed = details.get("changed_columns")
            for record in _extract_detail_records(details, remediation_id):
                row = {
                    "dataset": dataset,
                    "issue": label,
                    "remediation_id": remediation_id,
                }
                row.update(record)
                if issue_changed:
                    row["__changed_columns"] = list(issue_changed)
                detail_rows.append(row)
        else:
            summary_rows.append(
                {
                    "dataset": dataset,
                    "issue": label,
                    "timestamp": timestamp,
                    "impact": json.dumps(details, default=str, sort_keys=True),
                }
            )

    meta_keys = {"dataset", "issue", "remediation_id", "record_state", "__highlight__"}
    grouped_indices: Dict[str, Dict[str, List[int]]] = defaultdict(
        lambda: defaultdict(list)
    )
    remediation_columns: Dict[str, Set[str]] = defaultdict(set)
    for idx, row in enumerate(detail_rows):
        remediation_id = row.get("remediation_id")
        record_state = row.get("record_state")
        if remediation_id is None or record_state is None:
            continue
        grouped_indices[remediation_id][record_state].append(idx)
        changed_columns = row.get("__changed_columns")
        if isinstance(changed_columns, list):
            remediation_columns[remediation_id].update(
                str(column) for column in changed_columns if column
            )
        elif row.get("record_state") == "removed":
            remediation_columns[remediation_id].update(
                key
                for key in row.keys()
                if key not in meta_keys and row.get(key) is not None
            )

    def update_highlight(row_idx: int, columns: Iterable[str], rem_id: str) -> None:
        normalized = [col for col in columns if col is not None]
        existing_highlight = set(detail_rows[row_idx].get("__highlight__", []))
        existing_changed = set(detail_rows[row_idx].get("__changed_columns", []))
        if normalized:
            existing_highlight.update(normalized)
            existing_changed.update(normalized)
            remediation_columns[rem_id].update(normalized)
        detail_rows[row_idx]["__highlight__"] = sorted(existing_highlight)
        detail_rows[row_idx]["__changed_columns"] = sorted(existing_changed)

    for remediation_id, states in grouped_indices.items():
        data_columns: Set[str] = set()
        for row_indices in states.values():
            for row_idx in row_indices:
                data_columns.update(
                    key for key in detail_rows[row_idx].keys() if key not in meta_keys
                )
        if not data_columns:
            continue

        original_indices = states.get("original", [])
        result_indices = states.get("result", [])
        removed_indices = states.get("removed", [])
        added_indices = states.get("added", [])
        changed_set = remediation_columns.setdefault(remediation_id, set())

        if (
            original_indices
            and not result_indices
            and not removed_indices
            and not added_indices
        ):
            for idx in original_indices:
                update_highlight(idx, [], remediation_id)
            continue

        if original_indices and result_indices:
            original_highlights: Dict[int, Set[str]] = {
                idx: set() for idx in original_indices
            }
            diff_cache: Dict[Tuple[int, int], Set[str]] = {}

            def diff_columns(ref_idx: int, result_idx: int) -> Set[str]:
                key = (ref_idx, result_idx)
                if key not in diff_cache:
                    reference_row = detail_rows[ref_idx]
                    result_row = detail_rows[result_idx]
                    diff_cache[key] = {
                        column
                        for column in data_columns
                        if not _values_equal(
                            reference_row.get(column), result_row.get(column)
                        )
                    }
                return diff_cache[key]

            ordered_originals = list(original_indices)
            mapping: Dict[int, int] = {}
            for result_idx in result_indices:
                best_ref = min(
                    ordered_originals,
                    key=lambda idx: (
                        len(diff_columns(idx, result_idx)),
                        ordered_originals.index(idx),
                    ),
                )
                mapping[result_idx] = best_ref

            for result_idx, ref_idx in mapping.items():
                changed_cols = diff_columns(ref_idx, result_idx)
                if not changed_cols:
                    changed_cols = changed_set
                update_highlight(result_idx, changed_cols, remediation_id)
                original_highlights[ref_idx].update(changed_cols)

            for ref_idx, columns in original_highlights.items():
                if not columns:
                    columns = changed_set
                update_highlight(ref_idx, columns, remediation_id)
        else:
            for idx in result_indices:
                update_highlight(idx, changed_set or data_columns, remediation_id)
            for idx in original_indices:
                update_highlight(idx, changed_set or data_columns, remediation_id)

        for idx in removed_indices:
            update_highlight(idx, data_columns, remediation_id)
        for idx in added_indices:
            update_highlight(idx, data_columns, remediation_id)
        for idx in states.get("reference", []):
            update_highlight(idx, [], remediation_id)

    for row in detail_rows:
        row.setdefault("__highlight__", [])
        row.pop("__changed_columns", None)

    return pd.DataFrame(summary_rows), pd.DataFrame(detail_rows)


def _add_remediation_separators(df: pd.DataFrame) -> pd.DataFrame:
    """Insert blank separator rows between remediation groups for readability."""
    if df.empty or "remediation_id" not in df.columns:
        return df

    ordered_frames: List[pd.DataFrame] = []
    ordered_ids: List[str] = []

    for remediation_id in df["remediation_id"]:
        if remediation_id not in ordered_ids:
            ordered_ids.append(remediation_id)

    blank_template = {column: None for column in df.columns}
    if "__highlight__" in blank_template:
        blank_template["__highlight__"] = []

    for remediation_id in ordered_ids:
        group = df[df["remediation_id"] == remediation_id]
        blank_row = blank_template.copy()
        ordered_frames.append(pd.DataFrame([blank_row]))

        delimiter = f"───── {remediation_id} ─────"
        separator_row = {column: delimiter for column in df.columns}
        if "__highlight__" in separator_row:
            separator_row["__highlight__"] = []
        ordered_frames.append(pd.DataFrame([separator_row]))
        ordered_frames.append(pd.DataFrame([blank_row]))
        ordered_frames.append(group)

    ordered_frames.append(pd.DataFrame([blank_template.copy()]))

    return pd.concat(ordered_frames, ignore_index=True)


def _write_dataset_sheet(
    writer: pd.ExcelWriter, dataset: str, summary: Dict[str, Any]
) -> None:
    sheet_name = dataset[:31] or dataset
    row = 0
    constraint_sections: List[Tuple[str, pd.DataFrame]] = []

    info_rows = [
        {"Metric": "Status", "Value": summary.get("status")},
        {"Metric": "Validated At", "Value": summary.get("timestamp")},
        {"Metric": "Checks Run", "Value": summary.get("total_checks")},
        {"Metric": "Checks Failed", "Value": summary.get("failed_checks")},
        {"Metric": "Row Count", "Value": summary.get("row_count")},
        {"Metric": "Validation JSON", "Value": summary.get("result_path")},
    ]
    pd.DataFrame(info_rows).to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=row
    )
    row += len(info_rows) + 2

    pandas_types = summary.get("pandas_dtypes") or {}
    db_types = summary.get("db_column_types") or {}
    dtype_columns = sorted({*pandas_types.keys(), *db_types.keys()})
    dtype_rows = [
        {
            "column": column,
            "pandas_dtype": pandas_types.get(column),
            "db_type": db_types.get(column),
        }
        for column in dtype_columns
    ]
    row = _write_section(
        writer,
        sheet_name,
        row,
        "Column Types",
        pd.DataFrame(dtype_rows),
    )

    row = _write_section(
        writer,
        sheet_name,
        row,
        "Rule Checks",
        pd.DataFrame(summary.get("checks", [])),
    )

    unique_entry = summary.get("unique_constraint") or {}
    if unique_entry:
        formatted = unique_entry.copy()
        null_samples = formatted.pop("null_samples", None)
        columns = formatted.get("columns")
        if isinstance(columns, (list, tuple)):
            formatted["columns"] = ", ".join(map(str, columns))
        key_null = formatted.get("key_null_counts")
        if isinstance(key_null, dict):
            formatted["columns_with_nulls"] = ", ".join(
                f"{col}:{count}" for col, count in key_null.items()
            )
        if isinstance(formatted.get("nullable_columns"), list):
            formatted["nullable_columns"] = ", ".join(
                map(str, formatted["nullable_columns"])
            )
        for key in ("missing_sample", "duplicate_sample"):
            if key in formatted:
                formatted[key] = json.dumps(formatted[key], default=str)
        unique_df = pd.DataFrame([formatted])
        if isinstance(null_samples, dict):
            for column, records in null_samples.items():
                if not records:
                    continue
                sample_df = pd.DataFrame(records)
                if sample_df.empty:
                    continue
                sample_df.insert(0, "constraint_column", column)
                constraint_sections.append(
                    (f"Unique Null Samples — {column}", sample_df)
                )
    else:
        unique_df = pd.DataFrame()
    row = _write_section(writer, sheet_name, row, "Unique Constraint", unique_df)

    fk_entries = []
    reference_sections: List[Tuple[str, pd.DataFrame]] = []

    for entry in summary.get("foreign_keys", []):
        formatted = entry.copy()
        sample_records = formatted.pop("sample_records", None)
        null_samples = formatted.pop("null_samples", None)
        reference_samples = formatted.pop("reference_rows", None)
        if "reference_dataset" in formatted and "referenced_table" not in formatted:
            formatted["referenced_table"] = formatted.pop("reference_dataset")
        if isinstance(formatted.get("columns"), list):
            formatted["columns"] = ", ".join(map(str, formatted["columns"]))
        if isinstance(formatted.get("reference_columns"), list):
            formatted["reference_columns"] = ", ".join(
                map(str, formatted["reference_columns"])
            )
        fk_entries.append(formatted)
        if sample_records:
            sample_df = pd.DataFrame(sample_records).head(10)
            if not sample_df.empty:
                sample_df.insert(
                    0,
                    "constraint_columns",
                    formatted.get("columns"),
                )
                constraint_sections.append(
                    (
                        f"Foreign Key Failures — {formatted.get('columns')}",
                        sample_df,
                    )
                )
        if null_samples:
            sample_df = pd.DataFrame(null_samples).head(10)
            if not sample_df.empty:
                sample_df.insert(
                    0,
                    "constraint_columns",
                    formatted.get("columns"),
                )
                constraint_sections.append(
                    (
                        f"Foreign Key Null Samples — {formatted.get('columns')}",
                        sample_df,
                    )
                )
        if reference_samples:
            sample_df = pd.DataFrame(reference_samples).head(10)
            if not sample_df.empty:
                sample_df.insert(
                    0,
                    "constraint_columns",
                    formatted.get("columns"),
                )
                reference_sections.append(
                    (
                        f"Foreign Key Reference Rows — {formatted.get('columns')}",
                        sample_df,
                    )
                )
    row = _write_section(
        writer,
        sheet_name,
        row,
        "Foreign Key Checks",
        pd.DataFrame(fk_entries),
    )

    null_summary = summary.get("missing_values", {}) or {}
    null_rows = [
        {"column": column, "null_count": count}
        for column, count in null_summary.items()
    ]
    row = _write_section(
        writer,
        sheet_name,
        row,
        "Null Value Summary",
        pd.DataFrame(null_rows),
    )

    coverage_info = summary.get("version_season_coverage") or {}
    if coverage_info:

        def _format_list(values: Iterable[str], chunk_size: int = 15) -> str:
            items = [str(value) for value in values if value is not None]
            if not items:
                return ""
            groups = [
                ", ".join(items[idx : idx + chunk_size])
                for idx in range(0, len(items), chunk_size)
            ]
            return "\n".join(groups)

        coverage_rows: List[Dict[str, Any]] = []
        coverage_rows.append(
            {"Metric": "Present Count", "Value": coverage_info.get("present_count", 0)}
        )
        if coverage_info.get("missing_count") is not None:
            coverage_rows.append(
                {
                    "Metric": "Missing Count",
                    "Value": coverage_info.get("missing_count"),
                }
            )
        if coverage_info.get("unexpected_count") is not None:
            coverage_rows.append(
                {
                    "Metric": "Unexpected Count",
                    "Value": coverage_info.get("unexpected_count"),
                }
            )
        if coverage_info.get("present"):
            coverage_rows.append(
                {
                    "Metric": "Present Seasons",
                    "Value": _format_list(coverage_info.get("present", [])),
                }
            )
        if coverage_info.get("missing"):
            coverage_rows.append(
                {
                    "Metric": "Missing Seasons",
                    "Value": _format_list(coverage_info.get("missing", [])),
                }
            )
        if coverage_info.get("unexpected"):
            coverage_rows.append(
                {
                    "Metric": "Unexpected Seasons",
                    "Value": _format_list(coverage_info.get("unexpected", [])),
                }
            )
        if coverage_info.get("baseline"):
            coverage_rows.append(
                {
                    "Metric": "Baseline Reference",
                    "Value": coverage_info.get("baseline"),
                }
            )
        row = _write_section(
            writer,
            sheet_name,
            row,
            "Version Season Coverage",
            pd.DataFrame(coverage_rows),
        )

    for section_title, section_df in constraint_sections + reference_sections:
        row = _write_section(
            writer,
            sheet_name,
            row,
            section_title,
            section_df,
        )

    raw_issues = summary.get("issues", [])
    rem_summary_df, rem_detail_df = _expand_remediation_issues(raw_issues)
    row = _write_section(
        writer,
        sheet_name,
        row,
        "Remediation Events",
        rem_summary_df,
    )
    if not rem_detail_df.empty:
        reference_rows = rem_detail_df[rem_detail_df.get("record_state") == "reference"]
        detail_rows = rem_detail_df[rem_detail_df.get("record_state") != "reference"]
        if not detail_rows.empty:
            non_null_columns = [
                column
                for column in detail_rows.columns
                if column == "__highlight__" or detail_rows[column].notna().any()
            ]
            detail_rows = detail_rows[non_null_columns]
            detail_rows = _add_remediation_separators(detail_rows)
            row = _write_section(
                writer,
                sheet_name,
                row,
                "Remediation Details",
                detail_rows,
            )
        if not reference_rows.empty:
            reference_rows = reference_rows.copy()
            if "record_state" not in reference_rows.columns:
                reference_rows.insert(0, "record_state", "reference")
            else:
                ordered = ["record_state"] + [
                    column
                    for column in reference_rows.columns
                    if column != "record_state"
                    and (
                        column == "__highlight__"
                        or reference_rows[column].notna().any()
                    )
                ]
                reference_rows = reference_rows[ordered]
            non_null_columns = [
                column
                for column in reference_rows.columns
                if column == "__highlight__" or reference_rows[column].notna().any()
            ]
            reference_rows = reference_rows[non_null_columns]
            reference_rows = _add_remediation_separators(reference_rows)
            row = _write_section(
                writer,
                sheet_name,
                row,
                "Remediation Reference Records",
                reference_rows,
            )

    notes_list = summary.get("notes") or []
    if summary.get("status") == "passed" and not raw_issues and not null_rows:
        notes_list = notes_list or ["No data quality issues detected."]

    if notes_list:
        notes_df = pd.DataFrame([{"message": note} for note in notes_list])
        row = _write_section(writer, sheet_name, row, "Notes", notes_df)


def _write_metadata_summary_sheet(writer: pd.ExcelWriter) -> None:
    sheet_name = "Metadata Summary"
    row = 0

    upstream = _fetch_upstream_dataset_names()
    loaded = set(GLOBAL_DATASET_METADATA.keys())
    configured = set(GLOBAL_CONFIGURED_DATASETS)

    coverage_rows: List[Dict[str, Any]] = []
    coverage_map: Dict[str, Dict[str, Any]] = {}

    for dataset in sorted(loaded):
        coverage_map[dataset] = {"dataset": dataset, "status": "Loaded"}

    if upstream is None:
        note = _UPSTREAM_DATASETS_ERROR or "Unable to fetch upstream dataset list."
        coverage_rows.append({"dataset": "-", "status": "Unavailable", "note": note})
    else:
        for dataset in sorted(upstream):
            if dataset in loaded:
                coverage_map.setdefault(
                    dataset, {"dataset": dataset, "status": "Loaded"}
                )
            elif dataset in configured:
                coverage_map.setdefault(
                    dataset,
                    {
                        "dataset": dataset,
                        "status": "Configured but not loaded",
                    },
                )
            else:
                coverage_map.setdefault(
                    dataset,
                    {
                        "dataset": dataset,
                        "status": "Not configured",
                    },
                )
        for dataset in sorted(loaded - upstream):
            coverage_map[dataset]["status"] = "Loaded (not found upstream)"

    for dataset in sorted(configured - set(coverage_map.keys())):
        coverage_map.setdefault(
            dataset,
            {
                "dataset": dataset,
                "status": "Configured but not loaded",
            },
        )

    if coverage_map:
        coverage_rows.extend(coverage_map.values())

    if coverage_rows:
        row = _write_section(
            writer,
            sheet_name,
            row,
            "Upstream Dataset Coverage",
            pd.DataFrame(coverage_rows),
        )

    unexpected_rows: List[Dict[str, Any]] = []
    missing_rows: List[Dict[str, Any]] = []

    for dataset in sorted(GLOBAL_DATASET_METADATA.keys()):
        metadata = GLOBAL_DATASET_METADATA.get(dataset, {})
        table_name = metadata.get("table_name")
        observed = set(str(col) for col in metadata.get("observed", set()))
        db_columns = set(str(col) for col in metadata.get("db_columns", set()))
        auto_columns = set(str(col) for col in metadata.get("auto_columns", set()))
        observed_with_auto = observed | auto_columns
        unexpected = sorted(observed - db_columns)
        missing = sorted(db_columns - observed_with_auto)
        if unexpected:
            unexpected_rows.append(
                {
                    "dataset": dataset,
                    "table_name": table_name,
                    "columns": ", ".join(unexpected),
                }
            )
        if missing:
            missing_rows.append(
                {
                    "dataset": dataset,
                    "table_name": table_name,
                    "columns": ", ".join(missing),
                }
            )

    if unexpected_rows:
        row = _write_section(
            writer,
            sheet_name,
            row,
            "Unexpected Columns",
            pd.DataFrame(unexpected_rows),
        )

    if missing_rows:
        row = _write_section(
            writer,
            sheet_name,
            row,
            "Missing Columns",
            pd.DataFrame(missing_rows),
        )

    if row == 0:
        _write_section(
            writer,
            sheet_name,
            row,
            "Summary",
            pd.DataFrame([{"message": "No metadata discrepancies detected."}]),
        )


def _write_section(
    writer: pd.ExcelWriter,
    sheet_name: str,
    start_row: int,
    title: str,
    data: pd.DataFrame,
) -> int:
    from openpyxl.styles import (
        Alignment,
        PatternFill,
        Font,
        Border,
        Side,
    )  # Local import to keep optional dependency

    header_df = pd.DataFrame({"Section": [title]})
    header_df.to_excel(
        writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row
    )
    worksheet = writer.sheets[sheet_name]
    section_row_excel = start_row + 1
    section_cell = worksheet.cell(row=section_row_excel, column=1)
    section_cell.font = Font(bold=True, size=14)
    section_cell.alignment = Alignment(vertical="center")
    section_cell.border = Border(top=Side(style="medium"), bottom=Side(style="medium"))
    section_cell.fill = PatternFill(
        start_color="E8EEF7", end_color="E8EEF7", fill_type="solid"
    )
    start_row += 1

    description = SECTION_DESCRIPTIONS.get(title)
    if description is None:
        if title.startswith("Unique Null Samples"):
            description = "Example rows where the unique key columns are null."
        elif title.startswith("Foreign Key Failures"):
            description = "Sample rows that fail the foreign key constraint."
        elif title.startswith("Foreign Key Null Samples"):
            description = "Example rows where foreign key columns are null."
    if description:
        desc_df = pd.DataFrame({"Description": [description]})
        desc_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=False,
            startrow=start_row,
        )
        desc_row_excel = start_row + 1
        desc_cell = worksheet.cell(row=desc_row_excel, column=1)
        desc_cell.font = Font(italic=True)
        start_row += len(desc_df) + 1

    df = data.copy()
    highlight_map: Dict[int, Set[str]] = {}
    if df is None or df.empty:
        df = pd.DataFrame([{"note": "None"}])
    else:
        if title.startswith("Remediation") and "__highlight__" in df.columns:
            highlight_series = df["__highlight__"]
            for idx, value in highlight_series.items():
                if isinstance(value, (list, tuple, set)):
                    highlight_set = {str(item) for item in value if item}
                elif value:
                    highlight_set = {str(value)}
                else:
                    highlight_set = set()
                if highlight_set:
                    highlight_map[idx] = highlight_set
            df = df.drop(columns="__highlight__")
        for column in df.columns:
            series = df[column]
            if pd.api.types.is_datetime64tz_dtype(series):
                try:
                    df[column] = series.dt.tz_localize(None)
                except (TypeError, AttributeError):
                    pass
            elif pd.api.types.is_object_dtype(series):

                def _strip_timezone(value: Any) -> Any:
                    if isinstance(value, pd.Timestamp) and value.tz is not None:
                        return value.tz_localize(None)
                    return value

                try:
                    df[column] = series.apply(_strip_timezone)
                except Exception:  # pragma: no cover - defensive
                    df[column] = series

    header_row = start_row
    data_row_start = start_row + 1

    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=header_row)

    worksheet = writer.sheets[sheet_name]
    max_lengths: Dict[int, int] = {}
    header_font = Font(bold=True)
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    highlight_null_fill = PatternFill(
        start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
    )
    highlight_fail_fill = PatternFill(
        start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
    )
    highlight_change_fill = PatternFill(
        start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
    )
    header_fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )
    stripe_fill = PatternFill(
        start_color="F5F7FB", end_color="F5F7FB", fill_type="solid"
    )

    for col_idx, column in enumerate(df.columns, start=1):
        header_length = len(str(column)) + 2
        data_values = df[column].astype(str).tolist()
        max_len = max([header_length] + [len(value) + 2 for value in data_values])
        # Allow wider columns for verbose text, but clamp at a reasonable maximum.
        max_lengths[col_idx] = max(18, min(max_len, 80))
        column_letter = worksheet.cell(row=header_row + 1, column=col_idx).column_letter
        worksheet.column_dimensions[column_letter].width = max_lengths[col_idx]
        header_cell = worksheet.cell(row=header_row + 1, column=col_idx)
        header_cell.font = header_font
        header_cell.alignment = Alignment(wrap_text=True, vertical="top")
        header_cell.border = border_style
        header_cell.fill = header_fill
        for row_idx in range(header_row + 2, header_row + len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border_style
            df_row_idx = row_idx - (header_row + 2)
            if df_row_idx >= 0 and df_row_idx % 2 == 0:
                if not cell.fill or cell.fill.fill_type is None:
                    cell.fill = stripe_fill
            if title == "Remediation Details":
                highlight_columns = highlight_map.get(df_row_idx, set())
                if df.columns[col_idx - 1] in highlight_columns:
                    cell.font = Font(bold=True)
                    cell.fill = highlight_change_fill

    if title == "Unique Constraint":
        if "columns_with_nulls" in df.columns:
            col_idx = df.columns.get_loc("columns_with_nulls") + 1
            for offset, value in enumerate(df["columns_with_nulls"], start=0):
                if isinstance(value, str) and value.strip():
                    excel_row = data_row_start + offset + 1
                    for excel_col in range(1, len(df.columns) + 1):
                        worksheet.cell(
                            row=excel_row, column=excel_col
                        ).fill = highlight_null_fill
        if "status" in df.columns:
            for offset, value in enumerate(df["status"], start=0):
                if isinstance(value, str) and value.lower() != "passed":
                    excel_row = data_row_start + offset + 1
                    for excel_col in range(1, len(df.columns) + 1):
                        worksheet.cell(
                            row=excel_row, column=excel_col
                        ).fill = highlight_fail_fill

    if title == "Foreign Key Checks":
        if "status" in df.columns:
            for offset, value in enumerate(df["status"], start=0):
                if isinstance(value, str) and value.lower() != "passed":
                    excel_row = data_row_start + offset + 1
                    for excel_col in range(1, len(df.columns) + 1):
                        worksheet.cell(
                            row=excel_row, column=excel_col
                        ).fill = highlight_fail_fill
        if "null_count" in df.columns:
            for offset, value in enumerate(df["null_count"], start=0):
                if isinstance(value, (int, float)) and value > 0:
                    excel_row = data_row_start + offset + 1
                    for excel_col in range(1, len(df.columns) + 1):
                        worksheet.cell(
                            row=excel_row, column=excel_col
                        ).fill = highlight_null_fill
    if title == "Version Season Coverage":
        if "Metric" in df.columns and "Value" in df.columns:
            for offset, metric in enumerate(df["Metric"], start=0):
                value = df["Value"].iloc[offset]
                if (
                    isinstance(metric, str)
                    and metric.lower().startswith("missing")
                    and value not in (None, "", 0)
                ):
                    excel_row = data_row_start + offset + 1
                    for excel_col in range(1, len(df.columns) + 1):
                        worksheet.cell(
                            row=excel_row, column=excel_col
                        ).fill = highlight_fail_fill
    return start_row + len(df) + 2
